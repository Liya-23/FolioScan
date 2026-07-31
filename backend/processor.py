# .\backend\processor.py 

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Callable

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font

LogFn = Callable[[str], None]


JUNK_ROW_TERMS = {
    "proof_of_address",
    "proof of address",
    "application",
    "mailbody",
}

NOISE_TERMS = {
    "email",
    "receipt",
    "read",
    "application",
    "proof",
    "address",
    "copy",
    "scan",
    "scanned",
    "document",
    "doc",
}

TRACKING_RE = re.compile(
    r"\b(?:PTA|JHB|CTR|CPT|DBN|PE|BFN|HA|REF|ID|CASE)[-_./ ]?\d{2,}\b",
    re.IGNORECASE,
)
DOB_RE = re.compile(r"(?<!\d)(\d{1,2})[-/.](\d{1,2})[-/.](\d{2,4})(?!\d)")
EMAIL_RE = re.compile(r"\b[\w.+-]+@[\w-]+(?:\.[\w-]+)+\b")
ID_RE = re.compile(r"\b\d{8,}\b")
READ_RECEIPT_RE = re.compile(r"(^|[^a-z0-9])read[-_ ]?receipts?([^a-z0-9]|$)", re.IGNORECASE)
EMAIL_SENT_RE = re.compile(r"(^|[^a-z0-9])e?[-_ ]?mail([^a-z0-9]|$)", re.IGNORECASE)


OUTPUT_COLUMNS = [
    "FileName",
    "FirstName",
    "SecondName",
    "LastName",
    "DOB",
    "FileExtension",
    "Email Sent",
    "Read Receipts",
    "Email Sent Date",
    "Read Receipt receved date",
    "FilePath",
    "Open File",
    "FileSizeKB",
    "DateModified",
    "DateCreated",
    "DateScanned",
    "Notes",
]


def normalize_dob(day: str, month: str, year: str) -> str:
    year_value = int(year)
    if len(year) == 2:
        year_value += 1900 if year_value > 30 else 2000

    try:
        parsed = datetime(int(year_value), int(month), int(day))
    except ValueError:
        return "Unavailable"

    return parsed.strftime("%d/%m/%Y")


def should_skip_for_names(raw: str) -> bool:
    value = raw.lower().replace("-", "_")
    return any(term in value for term in JUNK_ROW_TERMS)


def parse_filename(raw_value: object) -> dict[str, str]:
    original = str(raw_value).strip()
    filename_path = Path(original)
    text = filename_path.stem
    text = EMAIL_RE.sub(" ", text)

    dob = "Unavailable"
    match = DOB_RE.search(text)
    if match:
        dob = normalize_dob(*match.groups())
        text = text[: match.start()] + " " + text[match.end() :]

    text = TRACKING_RE.sub(" ", text)
    text = ID_RE.sub(" ", text)
    text = re.sub(r"\.[A-Za-z0-9]{2,5}$", " ", text)
    text = re.sub(r"[_\-./(){}\[\],]+", " ", text)

    parts = []
    if not should_skip_for_names(original):
        for item in text.split():
            cleaned = re.sub(r"[^A-Za-z']", "", item).strip("'").lower()
            if not cleaned or cleaned in NOISE_TERMS or len(cleaned) == 1:
                continue
            parts.append(cleaned.title())

    if not parts:
        first_name, second_name, last_name = "", "", ""
    elif len(parts) == 1:
        first_name, second_name, last_name = parts[0], "", ""
    elif len(parts) == 2:
        first_name, second_name, last_name = parts[0], "", parts[1]
    else:
        first_name = parts[0]
        second_name = " ".join(parts[1:-1])
        last_name = parts[-1]

    return {
        "FirstName": first_name,
        "SecondName": second_name,
        "LastName": last_name,
        "DOB": dob,
    }


def format_timestamp(timestamp: float) -> str:
    return datetime.fromtimestamp(timestamp).strftime("%d-%m-%Y %H:%M")


def build_file_record(file_path: Path, date_scanned: str) -> dict[str, object]:
    stats = file_path.stat()
    parsed = parse_filename(file_path.name)
    filename = file_path.name

    return {
        "FileName": filename,
        "FirstName": parsed["FirstName"],
        "SecondName": parsed["SecondName"],
        "LastName": parsed["LastName"],
        "DOB": parsed["DOB"],
        "FileExtension": file_path.suffix.lower().lstrip("."),
        "Email Sent": "yes" if EMAIL_SENT_RE.search(filename) else "n/a",
        "Read Receipts": "available" if READ_RECEIPT_RE.search(filename) else "unavailable",
        "Email Sent Date": "",
        "Read Receipt receved date": "",
        "FilePath": str(file_path),
        "Open File": "📄 Open",
        "FileSizeKB": round(stats.st_size / 1024, 2),
        "DateModified": format_timestamp(stats.st_mtime),
        "DateCreated": format_timestamp(stats.st_ctime),
        "DateScanned": date_scanned,
        "Notes": "",
    }

def add_excel_hyperlinks(excel_path: Path):
    wb = load_workbook(excel_path)
    ws = wb.active

    headers = {cell.value: cell.column for cell in ws[1]}

    path_col = headers["FilePath"]
    open_col = headers["Open File"]

    for row in range(2, ws.max_row + 1):
        filepath = ws.cell(row=row, column=path_col).value

        if filepath:
            cell = ws.cell(row=row, column=open_col)
            cell.value = "📄 Open"
            cell.hyperlink = filepath
            cell.style = "Hyperlink"

    wb.save(excel_path)

def process_folder(folder_path: str, log: LogFn = print) -> dict:
    source = Path(folder_path).expanduser()
    if not source.exists():
        raise FileNotFoundError(f"Folder does not exist: {source}")
    if not source.is_dir():
        raise ValueError("Input path must be a folder/directory.")

    log(f"Scanning folder tree: {source}")
    records = []
    errors = 0
    date_scanned = datetime.now().strftime("%d-%m-%Y %H:%M")
    output_path = source / "ITPC2.xlsx"

    for index, file_path in enumerate((item for item in source.rglob("*") if item.is_file()), start=1):
        if file_path.resolve() == output_path.resolve():
            continue
        try:
            records.append(build_file_record(file_path, date_scanned))
        except OSError as exc:
            errors += 1
            log(f"Skipped unreadable file: {file_path} ({exc})")
            continue
        if index % 100 == 0:
            log(f"Scanned {index} files...")

    output_df = pd.DataFrame(records, columns=OUTPUT_COLUMNS)

    output_df.to_excel(output_path, index=False)
    add_excel_hyperlinks(output_path)
    log(f"Saved cleaned file: {output_path}")

    return {
        "output_path": str(output_path),
        "total": int(len(records) + errors),
        "saved": int(len(records)),
        "skipped": int(errors),
    }


process_excel = process_folder
